"""
XLSX document loader with intelligent processing.
"""

import logging
import warnings
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl
import pandas as pd

logger = logging.getLogger(__name__)


class XLSXLoader:
    """Loads and intelligently processes content from XLSX files."""

    def __init__(self):
        """Initialize the XLSX loader."""
        self.max_rows_per_sheet = 1000  # Limit for large sheets
        self.max_sheets_to_process = 20  # Limit for workbooks with many sheets
        self.max_columns_summary = 25  # Limit for very wide sheets

    def _detect_sheet_type(self, df: pd.DataFrame, sheet_name: str) -> Dict[str, Any]:
        """
        Detect the type and purpose of a worksheet.

        Args:
            df: Pandas DataFrame of the sheet
            sheet_name: Name of the worksheet

        Returns:
            Dictionary with sheet type analysis
        """
        analysis = {
            "type": "data",
            "has_formulas": False,
            "is_summary": False,
            "is_pivot": False,
            "is_dashboard": False,
            "data_density": 0,
            "structure": "tabular",
        }

        # Analyze sheet name for clues
        name_lower = sheet_name.lower()
        if any(
            keyword in name_lower
            for keyword in ["summary", "overview", "dashboard", "report"]
        ):
            analysis["is_summary"] = True
            analysis["type"] = "summary"
        elif any(
            keyword in name_lower for keyword in ["pivot", "analysis", "breakdown"]
        ):
            analysis["is_pivot"] = True
            analysis["type"] = "analysis"
        elif any(
            keyword in name_lower
            for keyword in ["raw", "data", "detail", "transaction"]
        ):
            analysis["type"] = "raw_data"

        # Calculate data density
        total_cells = df.shape[0] * df.shape[1]
        non_empty_cells = df.notna().sum().sum()
        analysis["data_density"] = (
            non_empty_cells / total_cells if total_cells > 0 else 0
        )

        # Detect structure patterns
        if df.shape[1] > df.shape[0] and df.shape[1] > 10:
            analysis["structure"] = "wide_format"
        elif analysis["data_density"] < 0.3 and df.shape[0] < 50:
            analysis["structure"] = (
                "sparse_layout"  # Likely a formatted report/dashboard
            )

        return analysis

    def _analyze_sheet_content(
        self, df: pd.DataFrame, sheet_name: str, workbook_context: Dict
    ) -> str:
        """
        Analyze and create intelligent summary of sheet content.

        Args:
            df: Pandas DataFrame
            sheet_name: Name of the worksheet
            workbook_context: Context about the entire workbook

        Returns:
            Formatted analysis text
        """
        analysis = []

        # Sheet header
        analysis.append(f"=== WORKSHEET: '{sheet_name}' ===")

        # Detect sheet type
        sheet_type_info = self._detect_sheet_type(df, sheet_name)
        analysis.append(
            f"Sheet type: {sheet_type_info['type'].replace('_', ' ').title()}"
        )
        analysis.append(f"Dimensions: {df.shape[0]} rows Ã— {df.shape[1]} columns")
        analysis.append(f"Data density: {sheet_type_info['data_density']:.1%}")
        analysis.append(f"Structure: {sheet_type_info['structure'].replace('_', ' ')}")
        analysis.append("")

        # Column analysis
        if df.shape[1] > 0:
            analysis.append("=== Column Structure ===")

            for i, col in enumerate(df.columns):
                if i >= self.max_columns_summary:
                    analysis.append(f"... and {len(df.columns) - i} more columns")
                    break

                # Handle unnamed columns (common in Excel)
                col_name = str(col) if col and str(col) != "nan" else f"Column_{i+1}"

                # Analyze column content
                non_null_count = df[col].notna().sum()
                _null_count = df[col].isna().sum()

                if non_null_count > 0:
                    # Detect data type
                    sample_values = df[col].dropna().head(5)
                    dtype_info = self._analyze_column_type(df[col])

                    analysis.append(f"'{col_name}':")
                    analysis.append(f"  - Type: {dtype_info['primary_type']}")
                    analysis.append(
                        f"  - Values: {non_null_count}/{df.shape[0]} filled ({non_null_count/df.shape[0]*100:.1f}%)"
                    )

                    if dtype_info["is_numeric"]:
                        stats = df[col].describe()
                        analysis.append(
                            f"  - Range: {stats['min']:.2f} to {stats['max']:.2f} (avg: {stats['mean']:.2f})"
                        )
                    elif dtype_info["is_categorical"]:
                        unique_count = df[col].nunique()
                        analysis.append(f"  - Unique values: {unique_count}")
                        if unique_count <= 10:
                            top_values = df[col].value_counts().head(3)
                            top_list = [f"'{k}' ({v})" for k, v in top_values.items()]
                            analysis.append(f"  - Top values: {', '.join(top_list)}")

                    # Sample values
                    sample_str = ", ".join(str(v) for v in sample_values.tolist())
                    if len(sample_str) > 80:
                        sample_str = sample_str[:80] + "..."
                    analysis.append(f"  - Sample: {sample_str}")
                else:
                    analysis.append(f"'{col_name}': [Empty column]")

                analysis.append("")

        # Data insights specific to sheet type
        analysis.extend(self._generate_sheet_insights(df, sheet_name, sheet_type_info))

        return "\n".join(analysis)

    def _analyze_column_type(self, series: pd.Series) -> Dict[str, Any]:
        """
        Analyze the data type and characteristics of a column.

        Args:
            series: Pandas Series (column data)

        Returns:
            Dictionary with type analysis
        """
        analysis = {
            "primary_type": "text",
            "is_numeric": False,
            "is_categorical": False,
            "is_datetime": False,
            "is_boolean": False,
            "pattern": None,
        }

        # Drop null values for analysis
        clean_series = series.dropna()

        if len(clean_series) == 0:
            return analysis

        # Check for numeric data
        if pd.api.types.is_numeric_dtype(clean_series):
            analysis["primary_type"] = "numeric"
            analysis["is_numeric"] = True

            # Check if it's likely an ID or categorical numeric
            unique_ratio = clean_series.nunique() / len(clean_series)
            if unique_ratio > 0.95:
                analysis["primary_type"] = "identifier"
            elif clean_series.dtype == "int64" and unique_ratio < 0.1:
                analysis["primary_type"] = "categorical_numeric"
                analysis["is_categorical"] = True

        # Check for datetime
        elif pd.api.types.is_datetime64_any_dtype(clean_series):
            analysis["primary_type"] = "datetime"
            analysis["is_datetime"] = True

        # Check for boolean
        elif pd.api.types.is_bool_dtype(clean_series):
            analysis["primary_type"] = "boolean"
            analysis["is_boolean"] = True

        # Text analysis
        else:
            analysis["primary_type"] = "text"

            # Check if categorical (limited unique values)
            unique_ratio = clean_series.nunique() / len(clean_series)
            if unique_ratio < 0.5 and clean_series.nunique() < 50:
                analysis["is_categorical"] = True
                analysis["primary_type"] = "categorical_text"

            # Try to detect patterns
            sample_values = clean_series.head(10).astype(str)

            # Email pattern
            if sample_values.str.contains("@", na=False).any():
                analysis["pattern"] = "email"
            # Phone pattern
            elif (
                sample_values.str.contains(r"[\d\-\(\)\s]+", na=False).all()
                and sample_values.str.len().between(7, 15).all()
            ):
                analysis["pattern"] = "phone"
            # URL pattern
            elif sample_values.str.contains("http|www", na=False).any():
                analysis["pattern"] = "url"

        return analysis

    def _generate_sheet_insights(
        self, df: pd.DataFrame, sheet_name: str, sheet_type_info: Dict
    ) -> List[str]:
        """
        Generate business and analytical insights for the sheet.

        Args:
            df: Pandas DataFrame
            sheet_name: Name of the worksheet
            sheet_type_info: Sheet type analysis

        Returns:
            List of insight strings
        """
        insights = []
        insights.append("=== Sheet Insights ===")

        # Business context detection
        business_indicators = self._detect_business_context(df)
        if business_indicators:
            insights.append("Business context detected:")
            insights.extend([f"  - {indicator}" for indicator in business_indicators])

        # Relationship analysis
        potential_keys = []
        potential_measures = []
        potential_dates = []

        for col in df.columns:
            col_analysis = self._analyze_column_type(df[col])
            col_name = str(col).lower()

            # Identify key columns
            if col_analysis["primary_type"] == "identifier" or any(
                keyword in col_name for keyword in ["id", "key", "code", "number"]
            ):
                potential_keys.append(str(col))

            # Identify measure columns
            elif col_analysis["is_numeric"] and any(
                keyword in col_name
                for keyword in [
                    "amount",
                    "total",
                    "sum",
                    "count",
                    "revenue",
                    "cost",
                    "price",
                    "value",
                ]
            ):
                potential_measures.append(str(col))

            # Identify date columns
            elif col_analysis["is_datetime"] or any(
                keyword in col_name
                for keyword in ["date", "time", "created", "updated"]
            ):
                potential_dates.append(str(col))

        # Report findings
        if potential_keys:
            insights.append(
                f"Potential key/identifier columns: {', '.join(potential_keys[:5])}"
            )

        if potential_measures:
            insights.append(
                f"Potential measure/metric columns: {', '.join(potential_measures[:5])}"
            )

        if potential_dates:
            insights.append(f"Date/time columns: {', '.join(potential_dates[:3])}")

        # Analysis suggestions
        analysis_suggestions = []

        if potential_dates and potential_measures:
            analysis_suggestions.append("Time series analysis possible")

        if potential_keys and len(potential_measures) > 1:
            analysis_suggestions.append("Dimensional analysis by key groupings")

        if len(potential_measures) > 1:
            analysis_suggestions.append("Cross-metric correlation analysis")

        if sheet_type_info["data_density"] > 0.8 and df.shape[0] > 100:
            analysis_suggestions.append("Suitable for statistical analysis")

        if analysis_suggestions:
            insights.append("Recommended analyses:")
            insights.extend(
                [f"  - {suggestion}" for suggestion in analysis_suggestions]
            )

        return insights

    def _detect_business_context(self, df: pd.DataFrame) -> List[str]:
        """
        Detect business context from column names and data patterns.

        Args:
            df: Pandas DataFrame

        Returns:
            List of detected business contexts
        """
        contexts = []

        # Convert all column names to lowercase for analysis
        col_names = [str(col).lower() for col in df.columns]
        col_text = " ".join(col_names)

        # Financial context
        if any(
            keyword in col_text
            for keyword in [
                "revenue",
                "profit",
                "cost",
                "expense",
                "budget",
                "financial",
            ]
        ):
            contexts.append("Financial data")

        # Sales context
        if any(
            keyword in col_text
            for keyword in [
                "sales",
                "customer",
                "order",
                "product",
                "quantity",
                "price",
            ]
        ):
            contexts.append("Sales/commerce data")

        # HR context
        if any(
            keyword in col_text
            for keyword in ["employee", "salary", "department", "hire", "performance"]
        ):
            contexts.append("Human resources data")

        # Marketing context
        if any(
            keyword in col_text
            for keyword in ["campaign", "lead", "conversion", "channel", "marketing"]
        ):
            contexts.append("Marketing data")

        # Operations context
        if any(
            keyword in col_text
            for keyword in [
                "inventory",
                "supply",
                "production",
                "manufacturing",
                "logistics",
            ]
        ):
            contexts.append("Operations data")

        # Analytics context
        if any(
            keyword in col_text
            for keyword in ["metric", "kpi", "score", "rate", "ratio", "percentage"]
        ):
            contexts.append("Analytics/KPI data")

        return contexts

    def _create_sample_data(self, df: pd.DataFrame, sheet_name: str) -> str:
        """
        Create formatted sample data from the sheet.

        Args:
            df: Pandas DataFrame
            sheet_name: Name of the worksheet

        Returns:
            Formatted sample data
        """
        sample_lines = []
        sample_lines.append(f"\n=== Sample Data from '{sheet_name}' ===")

        if df.empty:
            sample_lines.append("No data in this sheet.")
            return "\n".join(sample_lines)

        # Show first few rows
        sample_rows = min(8, len(df))
        sample_df = df.head(sample_rows)

        sample_lines.append(f"First {sample_rows} rows:")

        # Create readable format
        for idx in range(sample_rows):
            row_data = []
            for col in sample_df.columns:
                col_name = (
                    str(col) if col and str(col) != "nan" else f"Col_{len(row_data)+1}"
                )
                value = str(sample_df.iloc[idx, sample_df.columns.get_loc(col)])

                # Truncate long values
                if len(value) > 40:
                    value = value[:37] + "..."

                row_data.append(f"{col_name}: {value}")

            sample_lines.append(f"Row {idx + 1}: {' | '.join(row_data)}")

        if len(df) > sample_rows:
            sample_lines.append(f"... and {len(df) - sample_rows} more rows")

        return "\n".join(sample_lines)

    def _get_workbook_overview(self, workbook_path: Path) -> Dict[str, Any]:
        """
        Get overview of the entire workbook structure.

        Args:
            workbook_path: Path to the Excel file

        Returns:
            Dictionary with workbook overview
        """
        overview = {
            "sheet_names": [],
            "sheet_count": 0,
            "total_cells_with_data": 0,
            "has_formulas": False,
            "has_charts": False,
        }

        try:
            # Use openpyxl to get detailed workbook info
            wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=False)
            overview["sheet_names"] = wb.sheetnames
            overview["sheet_count"] = len(wb.sheetnames)

            # Check for formulas and other features
            for sheet_name in wb.sheetnames[:5]:  # Check first 5 sheets
                try:
                    ws = wb[sheet_name]

                    # Count non-empty cells
                    for row in ws.iter_rows(
                        max_row=min(100, ws.max_row), max_col=min(50, ws.max_column)
                    ):
                        for cell in row:
                            if cell.value is not None:
                                overview["total_cells_with_data"] += 1

                                # Check for formulas
                                if str(cell.value).startswith("="):
                                    overview["has_formulas"] = True

                except Exception as e:
                    logger.debug(f"Error analyzing sheet {sheet_name}: {e}")
                    continue

            wb.close()

        except Exception as e:
            logger.debug(f"Error getting workbook overview: {e}")
            # Fallback to pandas
            try:
                with pd.ExcelFile(workbook_path) as xls:
                    overview["sheet_names"] = xls.sheet_names
                    overview["sheet_count"] = len(xls.sheet_names)
            except Exception as e2:
                logger.debug(f"Pandas fallback also failed: {e2}")

        return overview

    def load(self, file_path: Path) -> Optional[str]:
        """
        Load and intelligently process content from an XLSX file.

        Args:
            file_path: Path to the XLSX file

        Returns:
            Processed text content or None if failed
        """
        try:
            # Suppress openpyxl warnings
            warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

            # Get workbook overview
            workbook_overview = self._get_workbook_overview(file_path)

            if workbook_overview["sheet_count"] == 0:
                logger.warning(f"No sheets found in XLSX: {file_path}")
                return None

            content_parts = []

            # Add workbook summary
            content_parts.append(f"=== EXCEL WORKBOOK ANALYSIS: {file_path.name} ===")
            content_parts.append(f"Total sheets: {workbook_overview['sheet_count']}")
            content_parts.append(
                f"Sheet names: {', '.join(workbook_overview['sheet_names'])}"
            )

            if workbook_overview["has_formulas"]:
                content_parts.append("Contains formulas: Yes")

            content_parts.append("")

            # Process each sheet
            sheets_processed = 0

            try:
                with pd.ExcelFile(file_path) as xls:
                    for sheet_name in xls.sheet_names:
                        if sheets_processed >= self.max_sheets_to_process:
                            content_parts.append(
                                f"... and {len(xls.sheet_names) - sheets_processed} more sheets (limit reached)"
                            )
                            break

                        try:
                            # Read sheet with reasonable limits
                            df = pd.read_excel(
                                xls,
                                sheet_name=sheet_name,
                                nrows=self.max_rows_per_sheet,
                                na_values=[
                                    "",
                                    "NULL",
                                    "null",
                                    "NA",
                                    "n/a",
                                    "#N/A",
                                    "#REF!",
                                    "#VALUE!",
                                ],
                            )

                            if df.empty:
                                content_parts.append(
                                    f"WORKSHEET '{sheet_name}': [Empty]"
                                )
                                content_parts.append("")
                                continue

                            # Clean column names
                            df.columns = [
                                (
                                    str(col).strip()
                                    if col and str(col) != "nan"
                                    else f"Column_{i+1}"
                                )
                                for i, col in enumerate(df.columns)
                            ]

                            # Analyze sheet content
                            sheet_analysis = self._analyze_sheet_content(
                                df, sheet_name, workbook_overview
                            )
                            content_parts.append(sheet_analysis)

                            # Add sample data
                            sample_data = self._create_sample_data(df, sheet_name)
                            content_parts.append(sample_data)
                            content_parts.append("")

                            sheets_processed += 1

                        except Exception as e:
                            logger.warning(
                                f"Error processing sheet '{sheet_name}': {e}"
                            )
                            content_parts.append(
                                f"WORKSHEET '{sheet_name}': [Error loading - {str(e)[:100]}]"
                            )
                            content_parts.append("")
                            continue

            except Exception as e:
                logger.error(f"Error reading Excel file: {e}")
                return None

            # Add final workbook insights
            content_parts.append(
                self._generate_workbook_insights(workbook_overview, sheets_processed)
            )

            full_content = "\n".join(content_parts)

            logger.info(
                f"Successfully loaded XLSX: {file_path} "
                f"({sheets_processed} sheets processed)"
            )

            return full_content

        except Exception as e:
            logger.error(f"Failed to load XLSX {file_path}: {e}")
            return None

    def _generate_workbook_insights(
        self, overview: Dict[str, Any], sheets_processed: int
    ) -> str:
        """
        Generate insights about the entire workbook.

        Args:
            overview: Workbook overview data
            sheets_processed: Number of sheets that were processed

        Returns:
            Workbook insights text
        """
        insights = []
        insights.append("=== WORKBOOK INSIGHTS ===")

        # Workbook complexity
        if overview["sheet_count"] > 10:
            insights.append("Workbook complexity: High (many worksheets)")
        elif overview["sheet_count"] > 5:
            insights.append("Workbook complexity: Medium")
        else:
            insights.append("Workbook complexity: Low (few worksheets)")

        # Usage patterns
        if overview["has_formulas"]:
            insights.append(
                "Usage pattern: Active calculation workbook (contains formulas)"
            )
        else:
            insights.append("Usage pattern: Data storage workbook (primarily data)")

        # Sheet naming analysis
        sheet_names = [name.lower() for name in overview["sheet_names"]]

        if any("summary" in name or "dashboard" in name for name in sheet_names):
            insights.append("Structure: Contains summary/dashboard sheets")

        if any("raw" in name or "data" in name for name in sheet_names):
            insights.append("Structure: Contains raw data sheets")

        # Data organization
        if len(set(sheet_names)) == len(sheet_names):
            insights.append("Organization: Well-structured (unique sheet names)")

        insights.append(
            f"Processing status: {sheets_processed}/{overview['sheet_count']} sheets analyzed"
        )

        return "\n".join(insights)
